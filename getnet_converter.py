#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lógica de conversión de liquidaciones Getnet (PDF) al Excel de importación
"GenImpExp_PagosEImputaciones". Sin dependencias de CLI ni de Streamlit para
poder reutilizarse desde ambos.
"""

import csv
import re
import unicodedata
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS = ["Cliente", "PLAN", "Importe", "NumPago", "SubnumPago", "VtoPago",
           "NameOnDoc", "FECHA", "NRODOC", "CUENTAIMP", "CCCI"]

NUM = r"-?[\d.,]+\.\d{2}"


class ConversionError(Exception):
    """Error esperable de negocio (PDF con formato inesperado, establecimiento
    sin configurar, etc.) - se muestra tal cual al usuario."""


def _norm(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.lower().strip()


def _to_float(s):
    return float(s.replace(",", ""))


def _to_amount_str(value):
    return f"{value:.2f}"


def _parse_date(s):
    return datetime.strptime(s, "%d/%m/%Y")


# --------------------------------------------------------------------------
# 1. Extraccion de un PDF
# --------------------------------------------------------------------------

def parse_pdf(pdf_file):
    """pdf_file: ruta (str/Path) o file-like (ej. BytesIO de un upload)."""
    with pdfplumber.open(pdf_file) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    data = {}

    m = re.search(r"N[ºo]\s*Liquidaci[oó]n\s+(\S+)\s+Rut\s+(\S+)", full_text)
    if not m:
        raise ConversionError("No se encontró 'Nº Liquidación / Rut' en el PDF")
    data["nro_liquidacion"] = m.group(1)
    data["rut"] = m.group(2)

    m = re.search(r"Fecha Presentaci[oó]n\s+(\d{2}/\d{2}/\d{4})", full_text)
    if not m:
        raise ConversionError("No se encontró 'Fecha Presentación' en el PDF")
    data["fecha_presentacion"] = m.group(1)

    m = re.search(r"Establecimiento\s+(\S+)", full_text)
    if not m:
        raise ConversionError("No se encontró 'Establecimiento' en el PDF")
    data["establecimiento"] = m.group(1)

    m = re.search(r"Total\s+((?:" + NUM + r"\s+){7}" + NUM + r")\s*\nAjustes", full_text)
    if not m:
        raise ConversionError("No se encontró la fila 'Total' de Información de totales")
    totales = [_to_float(x) for x in m.group(1).split()]
    # Importe Presentado, Importe Devuelto, Importe Neto Liquidado,
    # Servicio Financiero, Otros Servicios, I.V.A, Ajustes, Importe Pagar
    data["servicio_financiero"] = totales[3]
    data["otros_servicios"] = totales[4]
    data["iva_totales"] = totales[5]

    m = re.search(r"Ajustes\n(.*?)\nDetalle de Transferencia", full_text, re.S)
    if not m:
        raise ConversionError("No se encontró la sección 'Ajustes'")
    ajustes = []
    for line in m.group(1).splitlines():
        line = line.strip()
        if not line or line.startswith("Fecha") or line.startswith("Total"):
            continue
        mm = re.match(r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(" + NUM + r")$", line)
        if mm:
            ajustes.append({
                "fecha": mm.group(1),
                "descripcion": mm.group(2).strip(),
                "importe": _to_float(mm.group(3)),
            })
    data["ajustes"] = ajustes

    m = re.search(r"Detalle de Transferencia\n(.*?)(?:\nCondiciones aceptadas|\Z)",
                   full_text, re.S)
    if not m:
        raise ConversionError("No se encontró 'Detalle de Transferencia'")
    transferencias = []
    for line in m.group(1).splitlines():
        line = line.strip()
        if not line or line.startswith("Vencimiento"):
            continue
        mm = re.match(
            r"^(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(\d{5,})\s+(\d+)\s+(" + NUM + r")$",
            line,
        )
        if mm:
            transferencias.append({
                "vencimiento": mm.group(1),
                "banco": mm.group(2).strip(),
                "cuenta": mm.group(3),
                "referencia": mm.group(4),
                "importe": _to_float(mm.group(5)),
            })
    data["transferencias"] = transferencias

    return data


# --------------------------------------------------------------------------
# 2. Tabla de configuracion por establecimiento
# --------------------------------------------------------------------------

def load_establecimientos(csv_path):
    config = {}
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            config[str(row["Establecimiento"]).strip()] = {
                "Cliente": row["Cliente"].strip(),
                "CUENTAIMP": row["CUENTAIMP"].strip(),
                "CCCI": row["CCCI"].strip(),
            }
    return config


def resolve_fijos(establecimiento, cfg):
    if establecimiento not in cfg:
        raise ConversionError(
            f"El establecimiento {establecimiento} no está en la tabla de "
            f"configuración (Cliente/CUENTAIMP/CCCI)."
        )
    return cfg[establecimiento]


# --------------------------------------------------------------------------
# 3. Construccion de filas de salida (para UN pdf ya parseado)
# --------------------------------------------------------------------------

def build_rows(data, fijos):
    cliente = int(fijos["Cliente"])
    cuentaimp = int(fijos["CUENTAIMP"])
    ccci = fijos["CCCI"]
    fecha_pres = data["fecha_presentacion"]
    nrodoc = int(data["nro_liquidacion"])

    rows = []
    warnings = []

    retencion_total = 0.0
    iva_ajuste_total = 0.0
    for aj in data["ajustes"]:
        desc = _norm(aj["descripcion"])
        if "compensac" in desc:
            # Las "Compensación" son movimientos internos que se anulan entre
            # sí (ver Detalle de Ajustes del PDF) y deben ignorarse por
            # completo: ni se suman ni se restan de ningún PLAN.
            continue
        if re.search(r"\bretenc", desc):
            retencion_total += abs(aj["importe"])
        elif re.search(r"\biva\b", desc):
            iva_ajuste_total += abs(aj["importe"])
        else:
            warnings.append(
                f"Liquidación {nrodoc}: ajuste no reconocido "
                f"('{aj['descripcion']}', {aj['importe']}) - no incluido."
            )

    def base_row(plan, importe, vto_pago, name_on_doc="", num_pago="0"):
        return {
            "Cliente": cliente,
            "PLAN": plan,
            "Importe": _to_amount_str(importe),
            "NumPago": num_pago,
            "SubnumPago": 0,
            "VtoPago": _parse_date(vto_pago),
            "NameOnDoc": name_on_doc,
            "FECHA": fecha_pres,
            "NRODOC": nrodoc,
            "CUENTAIMP": cuentaimp,
            "CCCI": ccci,
        }

    if retencion_total:
        rows.append(base_row("12", retencion_total, fecha_pres))
    if iva_ajuste_total:
        rows.append(base_row("13", iva_ajuste_total, fecha_pres))

    if data["iva_totales"]:
        rows.append(base_row("14", data["iva_totales"], fecha_pres))

    servicio_otros = data["servicio_financiero"] + data["otros_servicios"]
    if servicio_otros:
        rows.append(base_row("15", servicio_otros, fecha_pres))

    for t in data["transferencias"]:
        rows.append(base_row(
            "44", t["importe"], t["vencimiento"],
            name_on_doc="SANTANDER", num_pago=t["referencia"],
        ))

    return rows, warnings


# --------------------------------------------------------------------------
# 4. Combinar varios PDFs en un solo lote de filas
# --------------------------------------------------------------------------

def process_batch(pdf_files, cfg, names=None):
    """pdf_files: lista de rutas o file-like objects.
    Devuelve (rows, warnings, resumen) sin pisar datos entre documentos:
    cada PDF se parsea y construye de forma totalmente independiente y sus
    filas se concatenan al final."""
    names = names or [getattr(f, "name", str(f)) for f in pdf_files]
    all_rows = []
    warnings = []
    resumen = []
    vistos = {}  # nrodoc -> nombre de archivo, para detectar duplicados

    for pdf_file, name in zip(pdf_files, names):
        try:
            data = parse_pdf(pdf_file)
            fijos = resolve_fijos(data["establecimiento"], cfg)
            rows, w = build_rows(data, fijos)
        except ConversionError as e:
            warnings.append(f"{name}: {e}")
            continue

        nrodoc = data["nro_liquidacion"]
        if nrodoc in vistos:
            warnings.append(
                f"{name}: la liquidación Nº {nrodoc} ya fue cargada desde "
                f"'{vistos[nrodoc]}'. Se agrega igual, revisá que no sea un duplicado."
            )
        vistos[nrodoc] = name

        all_rows.extend(rows)
        warnings.extend(w)
        resumen.append({
            "archivo": name,
            "nro_liquidacion": nrodoc,
            "establecimiento": data["establecimiento"],
            "filas": len(rows),
        })

    return all_rows, warnings, resumen


# --------------------------------------------------------------------------
# 5. Escritura del Excel
# --------------------------------------------------------------------------

def write_excel(rows, dest):
    """dest: ruta (str/Path) o file-like (ej. BytesIO para descarga)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GenImpExp_PagosEImputaciones"

    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col_name])
            if col_name == "VtoPago":
                cell.number_format = "mm-dd-yy"
            elif col_name in ("Importe", "NumPago", "FECHA", "CCCI", "NameOnDoc"):
                cell.number_format = "@"

    wb.save(dest)
